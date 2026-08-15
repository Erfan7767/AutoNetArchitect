"""Bilingual Excel report generator."""
from __future__ import annotations
from pathlib import Path
from typing import Any, Mapping, Sequence
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from ._common import metadata, sanitize
from .report_models import ReportArtifact, ReportLanguage

class ExcelGenerator:
    """Generate XLSX workbooks with metadata, data, assumptions, and SoT basis."""
    def generate(self, *, title: str, records: Sequence[Mapping[str, Any]], output_path: str | Path, language: ReportLanguage | str = ReportLanguage.BOTH, sot_basis: Mapping[str, str] | None = None, evidence_basis: Sequence[str] = (), assumptions: Sequence[str] = ()) -> ReportArtifact:
        """Generate a redacted workbook."""
        selected = ReportLanguage(language)
        meta = metadata(title=title, language=selected, sot_basis=sot_basis, evidence_basis=evidence_basis)
        target = Path(output_path); target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        metadata_sheet = workbook.active; metadata_sheet.title = "Metadata"
        metadata_rows = [("title", title), ("report_id", meta.report_id), ("language", selected.value), ("generated_at", meta.generated_at.isoformat()), ("sot_basis", str(meta.sot_basis or {"status": "not supplied"})), ("evidence_basis", str(meta.evidence_basis or ["none supplied"])), ("redaction_applied", True), ("secret_values_included", False)]
        for row in metadata_rows: metadata_sheet.append(list(row))
        metadata_sheet["A1"].font = Font(bold=True)
        data_sheet = workbook.create_sheet("Records")
        safe_records = [sanitize(dict(item)) for item in records]
        keys = sorted({key for item in safe_records if isinstance(item, dict) for key in item})
        data_sheet.append(keys)
        for key in keys: data_sheet.cell(row=1, column=keys.index(key) + 1).font = Font(bold=True)
        for item in safe_records: data_sheet.append([item.get(key) for key in keys])
        assumptions_sheet = workbook.create_sheet("Assumptions")
        assumptions_sheet.append(["assumption"])
        for item in assumptions: assumptions_sheet.append([str(item)])
        sot_sheet = workbook.create_sheet("SoT_Basis")
        sot_sheet.append(["domain", "record_or_status"])
        for key, value in (meta.sot_basis or {"status": "not supplied"}).items(): sot_sheet.append([key, value])
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row: cell.alignment = Alignment(vertical="top", wrap_text=True)
        workbook.save(target)
        return ReportArtifact(metadata=meta, output_path=str(target), format="xlsx")
