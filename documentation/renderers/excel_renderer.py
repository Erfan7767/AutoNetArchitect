"""Excel renderer for tables and editable engineering schedules."""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from ..formatters.table_formatter import TableFormatter


class ExcelRenderer:
    """Render each documentation section as an editable worksheet."""

    def render(self, content: dict[str, Any], output_path: str, *, watermark: str = "DRAFT") -> int:
        """Write a workbook with formatted sheets and a metadata sheet."""
        workbook = Workbook()
        default = workbook.active
        workbook.remove(default)
        metadata = workbook.create_sheet("Document")
        metadata.append(["Title", f"{content.get('title_en', '')} / {content.get('title_ar', '')}"])
        metadata.append(["Generated", content.get("generated_at", "PENDING")])
        metadata.append(["Completeness", content.get("completeness_score", 0)])
        metadata.append(["Watermark", watermark or "NONE"])
        for cell in metadata[1]:
            cell.font = Font(bold=True)
        for section in content.get("sections", []):
            name = str(section.get("section_id", "section"))[:31]
            sheet = workbook.create_sheet(name or "Section")
            sheet.freeze_panes = "A2"
            sheet.append([str(section.get("title_en", "")), str(section.get("title_ar", ""))])
            sheet.append(["Status", str(section.get("status", "pending"))])
            value = section.get("content")
            headers, rows = TableFormatter().normalize(value)
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for column in sheet.columns:
                width = max(12, min(45, max(len(str(cell.value or "")) for cell in column) + 2))
                sheet.column_dimensions[column[0].column_letter].width = width
            validation = DataValidation(type="list", formula1='"complete,partial,pending,not_applicable"', allow_blank=False)
            sheet.add_data_validation(validation)
            validation.add(sheet["B2"])
            sheet.conditional_formatting.add("B2", FormulaRule(formula=['B2="pending"'], fill=PatternFill("solid", fgColor="FFF2CC")))
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return max(1, len(content.get("sections", [])) + 1)
